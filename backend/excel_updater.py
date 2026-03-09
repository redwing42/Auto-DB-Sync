"""Excel updater for RedWing DB Automation.

Handles Steps 2–9 of the approval pipeline:
  - Resolve Network, Location, Landing Zone IDs
  - Register Waypoint File
  - Create Flight Route

Uses openpyxl to preserve formatting. Supports dry_run mode for resolve-preview.
Uses temp-copy pattern: writes to _temp.xlsx, then renames on success.
"""

from __future__ import annotations

import logging
import re
import shutil
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from models import (
    EntityAction,
    EntityPreview,
    ResolvePreviewResponse,
    SubmissionPayload,
)

logger = logging.getLogger(__name__)

# ── Sheet names ──────────────────────────────────────────────────────────────
SHEET_NETWORKS = "Networks"
SHEET_LOCATIONS = "Locations"
SHEET_LANDING_ZONES = "Landing Zone"
SHEET_WAYPOINT_FILES = "Waypoint Files"
SHEET_FLIGHT_ROUTES = "Flight Routes"


def _auto_generate_code(name: str, max_len: int = 8) -> str:
    """Generate a short uppercase code from a location name.

    E.g. 'HQ - Redwing Techworks' → 'HQ-RWT'
    """
    # Split on spaces, hyphens, and other separators
    parts = re.split(r"[\s\-–—/]+", name.strip())
    # Take uppercase first letter of each significant word (skip single-char connectors)
    initials = []
    for p in parts:
        p = p.strip()
        if p and len(p) > 0:
            initials.append(p[0].upper())

    code = "".join(initials)

    # If we have a dash-separated name like 'HQ - Redwing Techworks',
    # try to preserve the structure
    if " - " in name:
        left, right = name.split(" - ", 1)
        left_parts = re.split(r"\s+", left.strip())
        right_parts = re.split(r"\s+", right.strip())
        left_code = "".join(p[0].upper() for p in left_parts if p)
        right_code = "".join(p[0].upper() for p in right_parts if p)
        code = f"{left_code}-{right_code}"

    return code[:max_len]


def _get_true_max_row(ws: Worksheet) -> int:
    """Find the last row that actually contains data (openpyxl max_row can be misleading)."""
    for row in range(ws.max_row, 0, -1):
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=row, column=col).value is not None:
                return row
    return 1


def _get_last_id(ws: Worksheet) -> int:
    """Get the ID from the last data row (assumes column A is 'id')."""
    true_max = _get_true_max_row(ws)
    if true_max < 2:
        return 0
    val = ws.cell(row=true_max, column=1).value
    try:
        return int(val) if val is not None else 0
    except (TypeError, ValueError):
        return 0


def _find_row_by_column_value(
    ws: Worksheet, col_idx: int, value: Any
) -> Optional[int]:
    """Find the first row where column col_idx matches value (case-sensitive string match)."""
    true_max = _get_true_max_row(ws)
    for row in range(2, true_max + 1):
        cell_val = ws.cell(row=row, column=col_idx).value
        if cell_val is not None and str(cell_val).strip() == str(value).strip():
            return row
    return None


def _get_header_map(ws: Worksheet) -> Dict[str, int]:
    """Return {header_name: column_index} from the first row."""
    headers = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val:
            headers[str(val).strip()] = col
    return headers


class ExcelUpdater:
    """Reads / updates the flight data Excel workbook."""

    def __init__(self, excel_path: Path):
        self.excel_path = excel_path
        self.wb: Optional[openpyxl.Workbook] = None
        self._header_maps: Dict[str, Dict[str, int]] = {}

    def open(self) -> None:
        self.wb = openpyxl.load_workbook(str(self.excel_path))
        for sheet_name in [
            SHEET_NETWORKS,
            SHEET_LOCATIONS,
            SHEET_LANDING_ZONES,
            SHEET_WAYPOINT_FILES,
            SHEET_FLIGHT_ROUTES,
        ]:
            ws = self.wb[sheet_name]
            self._header_maps[sheet_name] = _get_header_map(ws)

    def close(self) -> None:
        if self.wb:
            self.wb.close()
            self.wb = None
            self._header_maps.clear()

    def _ws(self, name: str) -> Worksheet:
        assert self.wb is not None, "Workbook not open"
        return self.wb[name]

    def _headers(self, sheet: str) -> Dict[str, int]:
        return self._header_maps[sheet]

    # ── Step 2: Resolve Network ──────────────────────────────────────────

    def resolve_network(
        self, network_name: str, dry_run: bool = False
    ) -> Tuple[EntityPreview, Optional[int]]:
        """Look up network by name (supports fuzzy/partial matching)."""
        ws = self._ws(SHEET_NETWORKS)
        headers = self._headers(SHEET_NETWORKS)
        name_col = headers.get("name", 2)
        id_col = headers.get("id", 1)

        # Robust match: check for exact match first, then partial match
        true_max = _get_true_max_row(ws)
        search_name = network_name.strip().lower()
        
        found_row = None
        for r in range(2, true_max + 1):
            cell_val = str(ws.cell(row=r, column=name_col).value).strip().lower()
            if cell_val == search_name:
                found_row = r
                break
            if cell_val in search_name or search_name in cell_val:
                found_row = r
                # Don't break yet, exact match might come later
        
        if found_row is not None:
            nid = int(ws.cell(row=found_row, column=id_col).value)
            actual_name = str(ws.cell(row=found_row, column=name_col).value)
            return (
                EntityPreview(id=nid, name=actual_name, action=EntityAction.EXISTING),
                nid,
            )
        else:
            return (
                EntityPreview(name=network_name, action=EntityAction.NOT_FOUND),
                None,
            )

    # ── Step 3/5: Resolve Location ───────────────────────────────────────

    def resolve_location(
        self,
        location_name: str,
        network_id: int,
        district_id: int = 1,
        dry_run: bool = False,
    ) -> Tuple[EntityPreview, Optional[int]]:
        """Look up or create a location."""
        ws = self._ws(SHEET_LOCATIONS)
        headers = self._headers(SHEET_LOCATIONS)
        name_col = headers.get("name", 2)
        id_col = headers.get("id", 1)

        row = _find_row_by_column_value(ws, name_col, location_name)
        if row is not None:
            loc_id = int(ws.cell(row=row, column=id_col).value)
            return (
                EntityPreview(
                    id=loc_id, name=location_name, action=EntityAction.EXISTING
                ),
                loc_id,
            )

        # Need to create
        new_id = _get_last_id(ws) + 1
        if dry_run:
            return (
                EntityPreview(
                    id=new_id, name=location_name, action=EntityAction.NEW
                ),
                new_id,
            )

        # Write new row
        new_row = _get_true_max_row(ws) + 1
        code = _auto_generate_code(location_name)

        col_map = {
            "id": new_id,
            "name": location_name,
            "code": code,
            "district_id": district_id,
            "network_id": network_id,
            "location_type_id": 1,
            "landing_zone_count": 0,
        }
        for col_name, value in col_map.items():
            if col_name in headers:
                ws.cell(row=new_row, column=headers[col_name], value=value)

        logger.info(f"Created location: id={new_id}, name={location_name}, code={code}")
        return (
            EntityPreview(id=new_id, name=location_name, action=EntityAction.NEW),
            new_id,
        )

    # ── Step 4/6: Resolve Landing Zone ───────────────────────────────────

    def resolve_landing_zone(
        self,
        lz_name: str,
        location_id: int,
        latitude: float,
        longitude: float,
        dry_run: bool = False,
    ) -> Tuple[EntityPreview, Optional[int]]:
        """Look up or create a landing zone. Lat/long mismatch → new."""
        ws = self._ws(SHEET_LANDING_ZONES)
        headers = self._headers(SHEET_LANDING_ZONES)
        name_col = headers.get("name", 3)
        id_col = headers.get("id", 1)
        lat_col = headers.get("latitude", 4)
        lng_col = headers.get("longitude", 5)

        row = _find_row_by_column_value(ws, name_col, lz_name)
        if row is not None:
            # Check lat/long match
            existing_lat = ws.cell(row=row, column=lat_col).value
            existing_lng = ws.cell(row=row, column=lng_col).value
            try:
                lat_match = abs(float(existing_lat) - latitude) < 0.0001
                lng_match = abs(float(existing_lng) - longitude) < 0.0001
            except (TypeError, ValueError):
                lat_match = False
                lng_match = False

            if lat_match and lng_match:
                lz_id = int(ws.cell(row=row, column=id_col).value)
                return (
                    EntityPreview(
                        id=lz_id, name=lz_name, action=EntityAction.EXISTING
                    ),
                    lz_id,
                )
            # Mismatch — treat as new

        # Create new LZ
        new_id = _get_last_id(ws) + 1
        if dry_run:
            return (
                EntityPreview(id=new_id, name=lz_name, action=EntityAction.NEW),
                new_id,
            )

        new_row = _get_true_max_row(ws) + 1
        col_map = {
            "id": new_id,
            "location_id": location_id,
            "name": lz_name,
            "latitude": latitude,
            "longitude": longitude,
            "Status": 1,
            "altitude_pixhawk": 0,
            "altitude_srtm": 0,
            "altitude_google_earth": 0,
        }
        for col_name, value in col_map.items():
            if col_name in headers:
                ws.cell(row=new_row, column=headers[col_name], value=value)

        # Increment landing_zone_count in parent location
        self._increment_lz_count(location_id)

        logger.info(f"Created landing zone: id={new_id}, name={lz_name}")
        return (
            EntityPreview(id=new_id, name=lz_name, action=EntityAction.NEW),
            new_id,
        )

    def _increment_lz_count(self, location_id: int) -> None:
        """Increment the landing_zone_count for a location."""
        ws = self._ws(SHEET_LOCATIONS)
        headers = self._headers(SHEET_LOCATIONS)
        id_col = headers.get("id", 1)
        count_col = headers.get("landing_zone_count")
        if count_col is None:
            return

        row = _find_row_by_column_value(ws, id_col, location_id)
        if row is not None:
            current = ws.cell(row=row, column=count_col).value or 0
            ws.cell(row=row, column=count_col, value=int(current) + 1)

    # ── Step 7: Register Waypoint File ───────────────────────────────────

    def register_waypoint_file(
        self,
        mission_filename: str,
        drive_link: str,
        dry_run: bool = False,
    ) -> Tuple[EntityPreview, Optional[int]]:
        """Always create a new waypoint file entry."""
        ws = self._ws(SHEET_WAYPOINT_FILES)
        headers = self._headers(SHEET_WAYPOINT_FILES)
        new_id = _get_last_id(ws) + 1
        stem = Path(mission_filename).stem

        if dry_run:
            return (
                EntityPreview(
                    id=new_id, name=mission_filename, action=EntityAction.NEW
                ),
                new_id,
            )

        new_row = _get_true_max_row(ws) + 1
        col_map = {
            "id": new_id,
            "filename": mission_filename,
            "google_drive_filelink": drive_link,
            "local_filepath": f"./missions/{mission_filename}",
            "elevation_image": f"./Elevation and flight routes/{stem} elevation graph.png",
            "route_image": f"./Elevation and flight routes/{stem} flight route.png",
        }
        for col_name, value in col_map.items():
            if col_name in headers:
                ws.cell(row=new_row, column=headers[col_name], value=value)

        logger.info(f"Registered waypoint file: id={new_id}, filename={mission_filename}")
        return (
            EntityPreview(
                id=new_id, name=mission_filename, action=EntityAction.NEW
            ),
            new_id,
        )

    # ── Step 8: Create Flight Route ──────────────────────────────────────

    def create_flight_route(
        self,
        start_lz_id: int,
        end_lz_id: int,
        start_location_id: int,
        end_location_id: int,
        waypoint_file_id: int,
        network_id: int,
        takeoff_direction: int,
        approach_direction: int,
        dry_run: bool = False,
    ) -> Tuple[EntityPreview, Optional[int]]:
        """Create a new flight route entry."""
        ws = self._ws(SHEET_FLIGHT_ROUTES)
        headers = self._headers(SHEET_FLIGHT_ROUTES)
        new_id = _get_last_id(ws) + 1

        if dry_run:
            return (
                EntityPreview(id=new_id, action=EntityAction.NEW),
                new_id,
            )

        new_row = _get_true_max_row(ws) + 1
        col_map = {
            "id": new_id,
            "start_lz_id": start_lz_id,
            "end_lz_id": end_lz_id,
            "start_location_id": start_location_id,
            "end_location_id": end_location_id,
            "waypoint_file_id": waypoint_file_id,
            "network_id": network_id,
            "takeoff_direction": takeoff_direction,
            "approach_direction": approach_direction,
            "status": 1,
            "distance": None,
            "flight_duration": None,
        }
        for col_name, value in col_map.items():
            if col_name in headers:
                ws.cell(row=new_row, column=headers[col_name], value=value)

        logger.info(f"Created flight route: id={new_id}")
        return (
            EntityPreview(id=new_id, action=EntityAction.NEW),
            new_id,
        )

    # ── Save (with temp-copy pattern) ────────────────────────────────────

    def save(self) -> None:
        """Save workbook using temp-file pattern to prevent corruption."""
        temp_path = self.excel_path.with_name(
            self.excel_path.stem + "_temp" + self.excel_path.suffix
        )
        self.wb.save(str(temp_path))
        # Verify the temp file is valid
        try:
            verify_wb = openpyxl.load_workbook(str(temp_path))
            verify_wb.close()
        except Exception:
            temp_path.unlink(missing_ok=True)
            raise RuntimeError("Saved Excel file is corrupt — aborting save")
        # Replace original
        shutil.move(str(temp_path), str(self.excel_path))
        logger.info(f"Saved Excel: {self.excel_path}")

    # ── Full Resolve Preview (dry run) ───────────────────────────────────

    def resolve_preview(
        self, payload: SubmissionPayload
    ) -> ResolvePreviewResponse:
        """Dry-run of Steps 2–8. Returns what IDs would be resolved or created."""
        warnings: List[str] = []

        # Step 2 — Network
        net_preview, network_id = self.resolve_network(
            payload.network_name, dry_run=True
        )
        if net_preview.action == EntityAction.NOT_FOUND:
            warnings.append(
                "Network does not exist. Contact supervisor."
            )
            # Still continue dry run to show rest of info
            network_id = 0  # placeholder for downstream

        # Step 3 — Source Location
        src_loc_preview, src_loc_id = self.resolve_location(
            payload.source_location_name,
            network_id or 0,
            dry_run=True,
        )
        if src_loc_preview.action == EntityAction.NEW:
            warnings.append("Source location will be created")

        # Step 4 — Source Landing Zone
        src_lz_preview, src_lz_id = self.resolve_landing_zone(
            payload.source_takeoff_zone_name,
            src_loc_id or 0,
            payload.source_latitude,
            payload.source_longitude,
            dry_run=True,
        )
        if src_lz_preview.action == EntityAction.NEW:
            warnings.append("Source landing zone will be created")

        # Step 5 — Destination Location
        dst_loc_preview, dst_loc_id = self.resolve_location(
            payload.destination_location_name,
            network_id or 0,
            dry_run=True,
        )
        if dst_loc_preview.action == EntityAction.NEW:
            warnings.append("Destination location will be created")

        # Step 6 — Destination Landing Zone
        dst_lz_preview, dst_lz_id = self.resolve_landing_zone(
            payload.destination_landing_zone_name,
            dst_loc_id or 0,
            payload.destination_latitude,
            payload.destination_longitude,
            dry_run=True,
        )
        if dst_lz_preview.action == EntityAction.NEW:
            warnings.append("Destination landing zone will be created")

        # Count new LZs for summary warning
        new_lz_count = sum(
            1
            for p in [src_lz_preview, dst_lz_preview]
            if p.action == EntityAction.NEW
        )
        if new_lz_count > 0:
            warnings.append(f"{new_lz_count} new landing zone(s) will be created")

        # Step 7 — Waypoint file (always new)
        wp_preview, _ = self.register_waypoint_file(
            payload.mission_filename,
            payload.mission_drive_link,
            dry_run=True,
        )

        # Step 8 — Flight route (always new)
        fr_preview, _ = self.create_flight_route(
            start_lz_id=src_lz_id or 0,
            end_lz_id=dst_lz_id or 0,
            start_location_id=src_loc_id or 0,
            end_location_id=dst_loc_id or 0,
            waypoint_file_id=0,
            network_id=network_id or 0,
            takeoff_direction=payload.takeoff_direction,
            approach_direction=payload.approach_direction,
            dry_run=True,
        )

        return ResolvePreviewResponse(
            network=net_preview,
            source_location=src_loc_preview,
            source_lz=src_lz_preview,
            destination_location=dst_loc_preview,
            destination_lz=dst_lz_preview,
            waypoint_file=wp_preview,
            flight_route=fr_preview,
            warnings=warnings,
        )

    # ── Full Write Pipeline (Steps 2–9) ─────────────────────────────────

    def execute_pipeline(
        self, payload: SubmissionPayload
    ) -> Dict[str, Any]:
        """Run Steps 2–9, writing all rows. Returns dict of resolved IDs."""
        # Step 2
        _, network_id = self.resolve_network(payload.network_name)
        if network_id is None:
            raise RuntimeError(
                "Network not found. New networks require supervisor approval."
            )

        # Step 3
        _, src_loc_id = self.resolve_location(
            payload.source_location_name, network_id
        )

        # Step 4
        _, src_lz_id = self.resolve_landing_zone(
            payload.source_takeoff_zone_name,
            src_loc_id,
            payload.source_latitude,
            payload.source_longitude,
        )

        # Step 5
        _, dst_loc_id = self.resolve_location(
            payload.destination_location_name, network_id
        )

        # Step 6
        _, dst_lz_id = self.resolve_landing_zone(
            payload.destination_landing_zone_name,
            dst_loc_id,
            payload.destination_latitude,
            payload.destination_longitude,
        )

        # Step 7
        _, wp_file_id = self.register_waypoint_file(
            payload.mission_filename,
            payload.mission_drive_link,
        )

        # Step 8
        _, route_id = self.create_flight_route(
            start_lz_id=src_lz_id,
            end_lz_id=dst_lz_id,
            start_location_id=src_loc_id,
            end_location_id=dst_loc_id,
            waypoint_file_id=wp_file_id,
            network_id=network_id,
            takeoff_direction=payload.takeoff_direction,
            approach_direction=payload.approach_direction,
        )

        # Step 9
        self.save()

        return {
            "network_id": network_id,
            "source_location_id": src_loc_id,
            "source_lz_id": src_lz_id,
            "destination_location_id": dst_loc_id,
            "destination_lz_id": dst_lz_id,
            "waypoint_file_id": wp_file_id,
            "flight_route_id": route_id,
        }
    # ── Duplicate Detection ──────────────────────────────────────────────

    def is_duplicate_submission(self, payload: SubmissionPayload) -> bool:
        """Check if an identical flight route already exists in the Excel.
        
        Robust version: Match by Network, then search all routes for matching 
        Waypoint Filename + Directions + Coordinate Proximity.
        """
        # 1. Resolve network (names are usually stable, but support partial matches)
        preview, net_id = self.resolve_network(payload.network_name)
        if net_id is None:
            return False

        # 2. Search Flight Routes for match
        ws_fr = self._ws(SHEET_FLIGHT_ROUTES)
        h_fr = self._headers(SHEET_FLIGHT_ROUTES)
        ws_lz = self._ws(SHEET_LANDING_ZONES)
        h_lz = self._headers(SHEET_LANDING_ZONES)
        ws_wp = self._ws(SHEET_WAYPOINT_FILES)
        h_wp = self._headers(SHEET_WAYPOINT_FILES)
        
        true_max_fr = _get_true_max_row(ws_fr)
        
        # Cache LZ coords for the locations involved to speed up
        lz_coords = {} # id -> (lat, lon)
        
        def get_lz_coords(lz_id):
            if lz_id in lz_coords: return lz_coords[lz_id]
            # Find row in LZ sheet
            true_max_lz = _get_true_max_row(ws_lz)
            for r in range(2, true_max_lz + 1):
                cur_id = ws_lz.cell(row=r, column=h_lz["id"]).value
                if cur_id is not None and int(cur_id) == lz_id:
                    try:
                        lat = float(ws_lz.cell(row=r, column=h_lz["latitude"]).value)
                        lon = float(ws_lz.cell(row=r, column=h_lz["longitude"]).value)
                        lz_coords[lz_id] = (lat, lon)
                        return (lat, lon)
                    except (TypeError, ValueError):
                        break
            return None

        for r in range(2, true_max_fr + 1):
            if ws_fr.cell(row=r, column=h_fr["network_id"]).value != net_id:
                continue
            
            # Match directions and filenames first (cheapest)
            f_to = int(ws_fr.cell(row=r, column=h_fr["takeoff_direction"]).value)
            f_app = int(ws_fr.cell(row=r, column=h_fr["approach_direction"]).value)
            
            if (f_to != payload.takeoff_direction or f_app != payload.approach_direction):
                continue
                
            # Check waypoint filename
            wp_id_val = ws_fr.cell(row=r, column=h_fr["waypoint_file_id"]).value
            if wp_id_val is None: continue
            wp_id = int(wp_id_val)
            
            wp_row = _find_row_by_column_value(ws_wp, h_wp["id"], wp_id)
            if wp_row is None: continue
            
            existing_filename = str(ws_wp.cell(row=wp_row, column=h_wp["filename"]).value).strip()
            if existing_filename != payload.mission_filename.strip():
                continue
                
            # Finally, check coordinate proximity for start/end LZs
            start_lz_id = int(ws_fr.cell(row=r, column=h_fr["start_lz_id"]).value)
            end_lz_id = int(ws_fr.cell(row=r, column=h_fr["end_lz_id"]).value)
            
            start_coords = get_lz_coords(start_lz_id)
            end_coords = get_lz_coords(end_lz_id)
            
            if not start_coords or not end_coords: continue
            
            match_start = (abs(start_coords[0] - payload.source_latitude) < 0.0001 and 
                           abs(start_coords[1] - payload.source_longitude) < 0.0001)
            match_end = (abs(end_coords[0] - payload.destination_latitude) < 0.0001 and 
                         abs(end_coords[1] - payload.destination_longitude) < 0.0001)
                         
            if match_start and match_end:
                return True

        return False
